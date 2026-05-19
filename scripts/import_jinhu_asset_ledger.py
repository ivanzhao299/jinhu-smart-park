#!/usr/bin/env python3
"""Import Jinhu park asset ledger into biz asset tables.

The source workbook is a production-style floor ledger with merged cells.
This importer is intentionally idempotent: it upserts by building, floor, and
unit code and keeps unrelated existing rows untouched.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/mac/Downloads/园区房源台账5.15.xlsx")
DEFAULT_REPORT = Path("database/import-reports/jinhu_asset_ledger_20260518.json")
TENANT_ID = "10000001"
PARK_ID = "20000001"
PARK_CODE = "JH"
PARK_NAME = "金湖科创产业园"
IMPORT_USER = "codex-import"
SOURCE_NAME = "园区房源台账5.15.xlsx"


@dataclass(frozen=True)
class LedgerRow:
    source_row: int
    sequence: str
    building_code: str
    floor_label: str
    building_total_area: Decimal | None
    unit_area: Decimal
    ledger_status: str
    lease_text: str


@dataclass(frozen=True)
class BuildingRecord:
    building_code: str
    building_name: str
    floor_count: int
    build_area: Decimal
    sort_no: int
    remark: str


@dataclass(frozen=True)
class FloorRecord:
    floor_code: str
    building_code: str
    floor_no: int
    floor_name: str
    floor_area: Decimal
    sort_no: int
    remark: str


@dataclass(frozen=True)
class UnitRecord:
    unit_code: str
    building_code: str
    floor_code: str
    unit_name: str
    usage_type: int
    unit_area: Decimal
    use_area: Decimal
    rental_status: int
    fitting_status: int
    ref_price: Decimal
    status: int
    remark: str


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--apply", action="store_true", help="Write to the local PostgreSQL container")
    args = parser.parse_args()

    ledger_rows = read_ledger(args.workbook)
    buildings, floors, units = build_records(ledger_rows)
    report = build_report(args.workbook, ledger_rows, buildings, floors, units)
    write_report(args.report, report)

    print_summary(report)
    print(f"report: {args.report.resolve()}")
    if not args.apply:
        print("dry-run only; pass --apply to import")
        return

    sql = build_sql(buildings, floors, units)
    run_psql(sql)
    print("import applied")


def read_ledger(workbook: Path) -> list[LedgerRow]:
    wb = load_workbook(workbook, data_only=True, read_only=False)
    ws = wb["园区资产房源台账"]
    merged_values: dict[tuple[int, int], Any] = {}
    for cell_range in ws.merged_cells.ranges:
        value = ws.cell(cell_range.min_row, cell_range.min_col).value
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                merged_values[(row, col)] = value

    def cell(row: int, col: int) -> str:
        value = ws.cell(row, col).value
        if value is None:
            value = merged_values.get((row, col))
        return "" if value is None else str(value).strip()

    rows: list[LedgerRow] = []
    for source_row in range(4, ws.max_row + 1):
        building_code = cell(source_row, 2).upper()
        if not building_code:
            continue
        unit_area = parse_decimal(cell(source_row, 5))
        if unit_area is None:
            raise ValueError(f"row {source_row}: 单层面积不能为空")
        rows.append(
            LedgerRow(
                source_row=source_row,
                sequence=cell(source_row, 1),
                building_code=building_code,
                floor_label=cell(source_row, 3),
                building_total_area=parse_decimal(cell(source_row, 4)),
                unit_area=unit_area,
                ledger_status=cell(source_row, 6),
                lease_text=cell(source_row, 9),
            )
        )
    return rows


def build_records(rows: list[LedgerRow]) -> tuple[list[BuildingRecord], list[FloorRecord], list[UnitRecord]]:
    by_building: dict[str, list[LedgerRow]] = defaultdict(list)
    for row in rows:
        by_building[row.building_code].append(row)

    buildings: list[BuildingRecord] = []
    floors: list[FloorRecord] = []
    units: list[UnitRecord] = []

    for sort_no, building_code in enumerate(sorted(by_building, key=building_sort_key), 1):
        building_rows = by_building[building_code]
        expanded_floors = [
            (row, floor_code, floor_no, floor_display, floor_sort)
            for row in building_rows
            for floor_code, floor_no, floor_display, floor_sort in expand_floor_entries(row)
        ]
        explicit_total = next((row.building_total_area for row in building_rows if row.building_total_area is not None), None)
        build_area = explicit_total if explicit_total is not None else sum_decimal(row.unit_area for row in building_rows)
        source_rows = f"{building_rows[0].source_row}-{building_rows[-1].source_row}"
        buildings.append(
            BuildingRecord(
                building_code=building_code,
                building_name=f"{building_code}号楼",
                floor_count=len(expanded_floors),
                build_area=money(build_area),
                sort_no=sort_no,
                remark=f"生产台账导入：{SOURCE_NAME}；源行 {source_rows}",
            )
        )

        for row, floor_code, floor_no, floor_display, floor_sort in expanded_floors:
            remark = (
                f"生产台账导入：{SOURCE_NAME}；源行 {row.source_row}；"
                f"台账状态：{row.ledger_status or '未填'}；出租情况：{row.lease_text or '未填'}"
            )
            floors.append(
                FloorRecord(
                    floor_code=floor_code,
                    building_code=row.building_code,
                    floor_no=floor_no,
                    floor_name=f"{row.building_code} {floor_display}",
                    floor_area=money(row.unit_area),
                    sort_no=floor_sort,
                    remark=remark,
                )
            )
            units.append(
                UnitRecord(
                    unit_code=f"{floor_code}-U01",
                    building_code=row.building_code,
                    floor_code=floor_code,
                    unit_name=f"{row.building_code} {floor_display}",
                    usage_type=20,
                    unit_area=money(row.unit_area),
                    use_area=money(row.unit_area),
                    rental_status=map_rental_status(row.lease_text),
                    fitting_status=map_fitting_status(row.ledger_status),
                    ref_price=Decimal("0.00"),
                    status=1,
                    remark=remark,
                )
            )

    return buildings, floors, units


def build_report(
    workbook: Path,
    ledger_rows: list[LedgerRow],
    buildings: list[BuildingRecord],
    floors: list[FloorRecord],
    units: list[UnitRecord],
) -> dict[str, Any]:
    rental_counts = Counter(row.lease_text or "未填" for row in ledger_rows)
    status_counts = Counter(row.ledger_status or "未填" for row in ledger_rows)
    imported_rental_counts = Counter(unit.rental_status for unit in units)
    return {
        "source": str(workbook),
        "tenant_id": TENANT_ID,
        "park_id": PARK_ID,
        "park_code": PARK_CODE,
        "park_name": PARK_NAME,
        "ledger_rows": len(ledger_rows),
        "buildings": len(buildings),
        "floors": len(floors),
        "units": len(units),
        "park_total_area": str(money(sum_decimal(building.build_area for building in buildings))),
        "source_lease_counts": dict(rental_counts),
        "source_status_counts": dict(status_counts),
        "mapped_rental_status_counts": {str(key): value for key, value in imported_rental_counts.items()},
        "building_records": [serialize(record) for record in buildings],
    }


def build_sql(buildings: list[BuildingRecord], floors: list[FloorRecord], units: list[UnitRecord]) -> str:
    total_area = money(sum_decimal(building.build_area for building in buildings))
    sql_parts = [
        "BEGIN;",
        f"""
INSERT INTO biz_park (
  tenant_id, park_id, park_code, park_name, total_area, status, remark, create_by, update_by
) VALUES (
  {q(TENANT_ID)}, {q(PARK_ID)}, {q(PARK_CODE)}, {q(PARK_NAME)}, {num(total_area)}, 1,
  {q(f"生产台账导入：{SOURCE_NAME}；楼栋 {len(buildings)}；房源 {len(units)}")},
  {q(IMPORT_USER)}, {q(IMPORT_USER)}
)
ON CONFLICT (park_code) WHERE is_deleted = false
DO UPDATE SET
  park_name = EXCLUDED.park_name,
  total_area = EXCLUDED.total_area,
  status = EXCLUDED.status,
  remark = EXCLUDED.remark,
  update_by = EXCLUDED.update_by,
  update_time = now();
""",
        values_insert(
            table="biz_building",
            columns=["building_code", "building_name", "floor_count", "build_area", "sort_no", "remark"],
            rows=[
                [b.building_code, b.building_name, b.floor_count, b.build_area, b.sort_no, b.remark]
                for b in buildings
            ],
            select_sql=(
                "SELECT "
                f"{q(TENANT_ID)}, {q(PARK_ID)}, src.building_code, src.building_name, "
                "src.floor_count, src.build_area, 1, src.sort_no, "
                f"{q(IMPORT_USER)}, {q(IMPORT_USER)}, src.remark FROM src"
            ),
            insert_columns=[
                "tenant_id",
                "park_id",
                "building_code",
                "building_name",
                "floor_count",
                "build_area",
                "status",
                "sort_no",
                "create_by",
                "update_by",
                "remark",
            ],
            conflict="(building_code) WHERE is_deleted = false",
            update_columns=["building_name", "floor_count", "build_area", "status", "sort_no", "remark"],
        ),
        values_insert(
            table="biz_floor",
            columns=["floor_code", "building_code", "floor_no", "floor_name", "floor_area", "sort_no", "remark"],
            rows=[
                [f.floor_code, f.building_code, f.floor_no, f.floor_name, f.floor_area, f.sort_no, f.remark]
                for f in floors
            ],
            select_sql=(
                "SELECT "
                f"{q(TENANT_ID)}, {q(PARK_ID)}, b.id, src.floor_code, src.floor_no, src.floor_name, "
                "src.floor_area, 1, src.sort_no, "
                f"{q(IMPORT_USER)}, {q(IMPORT_USER)}, src.remark "
                "FROM src JOIN biz_building b ON b.building_code = src.building_code AND b.is_deleted = false"
            ),
            insert_columns=[
                "tenant_id",
                "park_id",
                "building_id",
                "floor_code",
                "floor_no",
                "floor_name",
                "floor_area",
                "status",
                "sort_no",
                "create_by",
                "update_by",
                "remark",
            ],
            conflict="(floor_code) WHERE is_deleted = false",
            update_columns=["building_id", "floor_no", "floor_name", "floor_area", "status", "sort_no", "remark"],
        ),
        values_insert(
            table="biz_unit",
            columns=[
                "unit_code",
                "building_code",
                "floor_code",
                "unit_name",
                "usage_type",
                "unit_area",
                "use_area",
                "rental_status",
                "fitting_status",
                "ref_price",
                "status",
                "remark",
            ],
            rows=[
                [
                    u.unit_code,
                    u.building_code,
                    u.floor_code,
                    u.unit_name,
                    u.usage_type,
                    u.unit_area,
                    u.use_area,
                    u.rental_status,
                    u.fitting_status,
                    u.ref_price,
                    u.status,
                    u.remark,
                ]
                for u in units
            ],
            select_sql=(
                "SELECT "
                f"{q(TENANT_ID)}, {q(PARK_ID)}, src.unit_code, src.unit_code, "
                "b.id, f.id, src.unit_name, src.usage_type, src.unit_area, src.use_area, "
                "src.rental_status, src.fitting_status, src.ref_price, src.status, "
                f"{q(IMPORT_USER)}, {q(IMPORT_USER)}, now(), {q(IMPORT_USER)}, src.remark "
                "FROM src "
                "JOIN biz_building b ON b.building_code = src.building_code AND b.is_deleted = false "
                "JOIN biz_floor f ON f.floor_code = src.floor_code AND f.is_deleted = false"
            ),
            insert_columns=[
                "tenant_id",
                "park_id",
                "unit_code",
                "code",
                "building_id",
                "floor_id",
                "unit_name",
                "usage_type",
                "unit_area",
                "use_area",
                "rental_status",
                "fitting_status",
                "ref_price",
                "status",
                "create_by",
                "update_by",
                "status_update_time",
                "status_update_by",
                "remark",
            ],
            conflict="(tenant_id, park_id, unit_code) WHERE is_deleted = false",
            update_columns=[
                "code",
                "building_id",
                "floor_id",
                "unit_name",
                "usage_type",
                "unit_area",
                "use_area",
                "rental_status",
                "fitting_status",
                "ref_price",
                "status",
                "status_update_time",
                "status_update_by",
                "remark",
            ],
        ),
        stale_cleanup("biz_unit", "unit_code", [unit.unit_code for unit in units]),
        stale_cleanup("biz_floor", "floor_code", [floor.floor_code for floor in floors]),
        stale_cleanup("biz_building", "building_code", [building.building_code for building in buildings]),
        "COMMIT;",
    ]
    return "\n".join(sql_parts)


def values_insert(
    *,
    table: str,
    columns: list[str],
    rows: list[list[Any]],
    select_sql: str,
    insert_columns: list[str],
    conflict: str,
    update_columns: list[str],
) -> str:
    values = ",\n".join("(" + ", ".join(sql_value(value) for value in row) + ")" for row in rows)
    updates = ",\n  ".join(
        [f"{column} = EXCLUDED.{column}" for column in update_columns]
        + ["update_by = EXCLUDED.update_by", "update_time = now()", "is_deleted = false"]
    )
    return f"""
WITH src({", ".join(columns)}) AS (
  VALUES
{values}
)
INSERT INTO {table} ({", ".join(insert_columns)})
{select_sql}
ON CONFLICT {conflict}
DO UPDATE SET
  {updates};
"""


def stale_cleanup(table: str, code_column: str, codes: list[str]) -> str:
    code_values = ", ".join(q(code) for code in sorted(set(codes)))
    return f"""
UPDATE {table}
   SET is_deleted = true,
       update_by = {q(IMPORT_USER + "-cleanup")},
       update_time = now()
 WHERE tenant_id = {q(TENANT_ID)}
   AND park_id = {q(PARK_ID)}
   AND create_by = {q(IMPORT_USER)}
   AND is_deleted = false
   AND {code_column} NOT IN ({code_values});
"""


def run_psql(sql: str) -> None:
    subprocess.run(
        ["docker", "exec", "-i", "jinhu-smart-park-postgres", "psql", "-U", "jinhu", "-d", "jinhu_smart_park", "-v", "ON_ERROR_STOP=1"],
        input=sql,
        text=True,
        check=True,
    )


def expand_floor_entries(row: LedgerRow) -> list[tuple[str, int, str, int]]:
    label = row.floor_label.strip().upper().replace(" ", "")
    if "/" not in label:
        return [normalize_floor(row.building_code, row.floor_label)]
    parts = [part for part in label.split("/") if part]
    if len(parts) < 2:
        return [normalize_floor(row.building_code, row.floor_label)]
    return [normalize_floor(row.building_code, part) for part in parts]


def normalize_floor(building_code: str, raw_label: str) -> tuple[str, int, str, int]:
    label = raw_label.strip().upper().replace(" ", "")
    if not label:
        return f"{building_code}-WHOLE", 0, "整栋", 0
    numbers = re.findall(r"\d+(?:\.\d+)?", label)
    if not numbers:
        code_suffix = re.sub(r"[^A-Z0-9_]+", "_", label).strip("_") or "FLOOR"
        return f"{building_code}-{code_suffix}", 0, raw_label, 0
    code_parts = [floor_code_part(number) for number in numbers]
    first = Decimal(numbers[0])
    floor_no = int(first * 10) if first % 1 else int(first)
    sort_no = floor_no
    suffix = code_parts[0] if len(code_parts) == 1 else f"{code_parts[0]}_{'_'.join(part.removeprefix('F') for part in code_parts[1:])}"
    return f"{building_code}-{suffix}", floor_no, raw_label, sort_no


def floor_code_part(number_text: str) -> str:
    number = Decimal(number_text)
    if number % 1:
        whole, fraction = number_text.split(".", 1)
        return f"F{int(whole):02d}_{fraction.rstrip('0')}"
    return f"F{int(number):02d}"


def map_rental_status(text: str) -> int:
    normalized = text.strip()
    if normalized == "空置":
        return 10
    if normalized == "已租":
        return 30
    if normalized == "已售":
        return 70
    if normalized in {"招商中心", "自用"}:
        return 60
    return 10


def map_fitting_status(text: str) -> int:
    normalized = text.strip()
    if "精装" in normalized:
        return 30
    if any(token in normalized for token in ["基础", "地板", "固化", "简装"]):
        return 20
    return 10


def building_sort_key(code: str) -> tuple[str, int, str]:
    match = re.match(r"([A-Z]+)(\d+)", code)
    if not match:
        return code, 0, code
    return match.group(1), int(match.group(2)), code


def parse_decimal(value: str) -> Decimal | None:
    text = value.strip().replace(",", "")
    if not text:
        return None
    return money(Decimal(text))


def sum_decimal(values: Any) -> Decimal:
    total = Decimal("0")
    for value in values:
        total += value
    return total


def money(value: Decimal) -> Decimal:
    return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def sql_value(value: Any) -> str:
    if isinstance(value, Decimal):
        return num(value)
    if isinstance(value, int):
        return str(value)
    return q(str(value))


def num(value: Decimal) -> str:
    return format(value, "f")


def q(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def serialize(record: Any) -> dict[str, Any]:
    data = asdict(record)
    return {key: (str(value) if isinstance(value, Decimal) else value) for key, value in data.items()}


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def print_summary(report: dict[str, Any]) -> None:
    print(
        "parsed: "
        f"{report['ledger_rows']} ledger rows, "
        f"{report['buildings']} buildings, "
        f"{report['floors']} floors, "
        f"{report['units']} units, "
        f"park total area {report['park_total_area']}"
    )
    print(f"source lease counts: {report['source_lease_counts']}")
    print(f"mapped rental status counts: {report['mapped_rental_status_counts']}")


if __name__ == "__main__":
    main()
